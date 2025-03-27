import openpyxl
import os
import re
import requests
import webbrowser
import threading
import shutil
from tkinter import END, Toplevel, Label, Button, messagebox, filedialog

GITHUB_REPO = "Maksymilianx/Excel_word_changer"
FALLBACK_VERSION = "1.2.0"

def fetch_latest_version():
    """Fetch the latest release tag from GitHub."""
    url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
    try:
        response = requests.get(url, timeout=5)
        response.raise_for_status()
        return response.json().get("tag_name", FALLBACK_VERSION)
    except requests.RequestException:
        return FALLBACK_VERSION

VERSION = fetch_latest_version()

def clean_pipes(text):
    """
    Process the input text line-by-line:
    - Collapse sequences like "|   |" into a single pipe.
    - Remove any leading pipe at the beginning of each line.
    - If a line ends with a pipe and the next line starts with a pipe, remove the extra leading pipe.
    """
    lines = text.splitlines()
    new_lines = []
    for line in lines:
        line = re.sub(r'\|\s*\|', '|', line)
        if line.startswith('|'):
            line = line[1:]
        new_lines.append(line)
    for i in range(len(new_lines) - 1):
        if new_lines[i].endswith('|') and new_lines[i+1].startswith('|'):
            new_lines[i+1] = new_lines[i+1][1:]
    return "\n".join(new_lines)

def remove_key_value_pair_from_cell(cell_value, key):
    pattern = r'\|?' + re.escape(key) + r'=[^|]*\|?'
    new_value = re.sub(pattern, '|', cell_value)
    new_value = clean_pipes(new_value).strip('|')
    return new_value if new_value != cell_value else None

def search_replace_or_remove_key(file_path, key, new_value, remove_key, log_widget, key_found):
    try:
        workbook = openpyxl.load_workbook(file_path)
        found = False
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        original_value = cell.value
                        if remove_key:
                            updated_cell = remove_key_value_pair_from_cell(cell.value, key)
                        else:
                            pattern = rf"\|?{key}=[^|]*\|?"
                            updated_cell = re.sub(pattern, f"|{key}={new_value}|", cell.value)
                            if updated_cell:
                                updated_cell = clean_pipes(updated_cell)
                        if updated_cell and updated_cell != original_value:
                            found = True
                            key_found[0] = True
                            cell.value = updated_cell
        if found:
            workbook.save(file_path)
            log_widget.insert(END, f"✅ Updated: {file_path}\n", "success")
    except Exception as e:
        log_widget.insert(END, f"❌ Error processing {file_path}: {e}\n", "error")

def backup_excel_files(source_dir, backup_dir, log_widget):
    """
    Copy all .xlsx files from source_dir (including subfolders) to backup_dir,
    preserving the folder structure. Create the backup directory if it doesn't exist.
    """
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir, exist_ok=True)
        log_widget.insert(END, f"ℹ Created backup directory: {backup_dir}\n", "info")
    else:
        log_widget.insert(END, f"ℹ Using existing backup directory: {backup_dir}\n", "info")
    for root, _, files in os.walk(source_dir):
        # Exclude the backup folder (compare absolute paths)
        if os.path.abspath(root) == os.path.abspath(backup_dir):
            continue
        for file in files:
            if file.endswith('.xlsx'):
                source_file = os.path.join(root, file)
                rel_path = os.path.relpath(root, source_dir)
                target_folder = os.path.join(backup_dir, rel_path)
                os.makedirs(target_folder, exist_ok=True)
                target_file = os.path.join(target_folder, file)
                shutil.copy2(source_file, target_file)
                log_widget.insert(END, f"Backup: {source_file} -> {target_file}\n", "info")

def process_excel_files(directory_path, backup_dir, key, new_value, remove_key, log_widget, progress_bar, percent_label):
    log_widget.delete(1.0, END)
    log_widget.insert(END, f"🔄 Processing files in {directory_path}...\n", "info")
    total_files = 0
    for root, dirs, files in os.walk(directory_path):
        if backup_dir:
            dirs[:] = [d for d in dirs if os.path.abspath(os.path.join(root, d)) != os.path.abspath(backup_dir)]
        for file in files:
            if file.endswith('.xlsx'):
                total_files += 1
    if total_files == 0:
        log_widget.insert(END, "❌ No Excel files found.\n", "error")
        progress_bar["value"] = 0
        percent_label.config(text="")
        return
    progress_bar["maximum"] = total_files
    progress_bar["value"] = 0
    key_found = [False]
    processed = 0
    progress_bar.grid()
    percent_label.grid()
    try:
        for root, dirs, files in os.walk(directory_path):
            if backup_dir:
                dirs[:] = [d for d in dirs if os.path.abspath(os.path.join(root, d)) != os.path.abspath(backup_dir)]
            for file in files:
                if file.endswith('.xlsx'):
                    file_path = os.path.join(root, file)
                    search_replace_or_remove_key(file_path, key, new_value, remove_key, log_widget, key_found)
                    processed += 1
                    progress_bar["value"] = processed
                    percent = int((processed / total_files) * 100)
                    percent_label.config(text=f"{percent}%")
                    progress_bar.update_idletasks()
        if not key_found[0]:
            log_widget.insert(END, f"⚠ Warning: The key '{key}' was not found in any file.\n", "warning")
            show_custom_warning_popup(f"The key '{key}' was not found in any file.")
        log_widget.insert(END, "✅ Process completed!\n", "success")
    except Exception as e:
        log_widget.insert(END, f"❌ An error occurred: {e}\n", "error")

def process_value_cells(file_path, old_value, new_value, log_widget):
    try:
        workbook = openpyxl.load_workbook(file_path)
        updated = False
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and old_value in cell.value:
                        cell.value = cell.value.replace(old_value, new_value)
                        updated = True
        if updated:
            workbook.save(file_path)
            log_widget.insert(END, f"✅ Processed cells replacing '{old_value}' with '{new_value}' in: {file_path}\n", "success")
        else:
            log_widget.insert(END, f"⚠ No cells containing '{old_value}' found in: {file_path}\n", "warning")
    except Exception as e:
        log_widget.insert(END, f"❌ Error processing cells in {file_path}: {e}\n", "error")

def process_value_in_directory(directory_path, old_value, new_value, log_widget, progress_bar, percent_label, backup_dir=None):
    total_files = 0
    for root, dirs, files in os.walk(directory_path):
        if backup_dir:
            dirs[:] = [d for d in dirs if os.path.abspath(os.path.join(root, d)) != os.path.abspath(backup_dir)]
        for file in files:
            if file.endswith('.xlsx'):
                total_files += 1
    if total_files == 0:
        log_widget.insert(END, "❌ No Excel files found.\n", "error")
        progress_bar["value"] = 0
        percent_label.config(text="")
        return
    progress_bar["maximum"] = total_files
    progress_bar["value"] = 0
    processed = 0
    progress_bar.grid()
    percent_label.grid()
    for root, dirs, files in os.walk(directory_path):
        if backup_dir:
            dirs[:] = [d for d in dirs if os.path.abspath(os.path.join(root, d)) != os.path.abspath(backup_dir)]
        for file in files:
            if file.endswith('.xlsx'):
                file_path = os.path.join(root, file)
                process_value_cells(file_path, old_value, new_value, log_widget)
                processed += 1
                progress_bar["value"] = processed
                percent = int((processed / total_files) * 100)
                percent_label.config(text=f"{percent}%")
                progress_bar.update_idletasks()
    log_widget.insert(END, "✅ Value replacement completed!\n", "success")

def start_value_replacement(directory, old_value, new_value, log_widget, progress_bar, percent_label, backup_dir_value):
    """
    For the Value Replacer: Validate the directory, determine the backup directory,
    perform backup, and then process value replacement.
    """
    if not os.path.exists(directory):
        log_widget.insert(END, "❌ Please select a valid processing directory.\n", "error")
        return
    backup_dir = backup_dir_value
    if not backup_dir:
        backup_dir = os.path.join(directory, "Backup")
        log_widget.insert(END, "ℹ No backup directory specified in Settings. Using default 'Backup' folder inside the processing directory.\n", "info")
    elif os.path.abspath(backup_dir) == os.path.abspath(directory):
        log_widget.insert(END, "⚠ Backup directory is the same as processing directory. Using a subfolder 'Backup' instead.\n", "warning")
        backup_dir = os.path.join(directory, "Backup")
    log_widget.insert(END, "🔄 Backing up files...\n", "info")
    backup_excel_files(directory, backup_dir, log_widget)
    log_widget.insert(END, "✅ Backup completed!\n", "success")
    progress_bar.grid()
    percent_label.grid()
    process_value_in_directory(directory, old_value, new_value, log_widget, progress_bar, percent_label, backup_dir)

def show_custom_warning_popup(message):
    popup = Toplevel()
    popup.title("Warning")
    popup.geometry("300x100")
    popup.update_idletasks()
    x_position = (popup.winfo_screenwidth() // 2) - (300 // 2)
    y_position = (popup.winfo_screenheight() // 2) - (100 // 2)
    popup.geometry(f"300x100+{x_position}+{y_position}")
    label = Label(popup, text=message, fg="red", font=("Arial", 12, "bold"))
    label.pack(pady=10)
    close_button = Button(popup, text="OK", command=popup.destroy)
    close_button.pack(pady=5)
    popup.grab_set()

def check_for_updates():
    latest_version = fetch_latest_version()
    if latest_version > VERSION:
        messagebox.showinfo("Update Available", f"A new version ({latest_version}) is available!\nVisit GitHub to download.")
    else:
        messagebox.showinfo("Up to Date", "You are using the latest version.")

def browse_directory(entry_widget):
    directory = filedialog.askdirectory()
    if directory:
        entry_widget.delete(0, END)
        entry_widget.insert(0, directory)

def open_github_link():
    webbrowser.open("https://github.com/Maksymilianx/Excel_word_changer")

def start_processing(directory_entry, key_entry, value_entry, remove_key_var, log_widget, progress_bar, percent_label, backup_dir_value):
    directory = directory_entry.get()
    key = key_entry.get()
    new_value = value_entry.get()
    remove_key = remove_key_var.get()
    backup_dir = backup_dir_value  # Passed in as parameter
    if not directory or not os.path.exists(directory):
        log_widget.insert(END, "❌ Please select a valid processing directory.\n", "error")
        return
    if not backup_dir:
        backup_dir = os.path.join(directory, "Backup")
        log_widget.insert(END, "ℹ No backup directory specified in Settings. Using default 'Backup' folder inside the processing directory.\n", "info")
    elif os.path.abspath(backup_dir) == os.path.abspath(directory):
        log_widget.insert(END, "⚠ Backup directory is the same as processing directory. Using a subfolder 'Backup' instead.\n", "warning")
        backup_dir = os.path.join(directory, "Backup")
    if not key:
        log_widget.insert(END, "❌ Please enter a key to search for.\n", "error")
        return
    if not remove_key and not new_value:
        log_widget.insert(END, "❌ Please enter a new value or check the 'Remove Key' option.\n", "error")
        return
    if remove_key:
        confirm = messagebox.askyesno("Confirm Deletion", f"Are you sure you want to remove '{key}' and its value?")
        if not confirm:
            log_widget.insert(END, "⚠ Deletion canceled by user.\n", "warning")
            return
    log_widget.insert(END, "🔄 Backing up files...\n", "info")
    backup_excel_files(directory, backup_dir, log_widget)
    log_widget.insert(END, "✅ Backup completed!\n", "success")
    progress_bar.grid()
    percent_label.grid()
    threading.Thread(target=process_excel_files, args=(directory, backup_dir, key, new_value, remove_key, log_widget, progress_bar, percent_label)).start()

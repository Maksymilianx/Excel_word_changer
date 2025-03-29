# test_functions.py
import os
import tempfile
import shutil
import pytest
import openpyxl
from functions import (
    fetch_latest_version,
    clean_pipes,
    remove_key_value_pair_from_cell,
    search_replace_or_remove_key,
    backup_excel_files,
    process_excel_files,
    process_value_cells,
    process_value_in_directory,
    start_value_replacement,
    check_for_updates,
    VERSION
)
import requests

# Dummy log widget to capture log messages.
class DummyLog:
    """A dummy log widget that stores log messages in a list."""
    def __init__(self):
        self.messages = []
    def insert(self, index, text, tag=None):
        self.messages.append(text)
    def delete(self, start, end):
        self.messages = []

# Dummy progress bar that supports item assignment.
class DummyProgressBar:
    def __init__(self):
        self.data = {}
    def __setitem__(self, key, value):
        self.data[key] = value
    def __getitem__(self, key):
        return self.data.get(key, None)
    def grid(self):
        pass
    def grid_remove(self):
        pass
    def update_idletasks(self):
        pass

# Dummy label for percentage display.
class DummyLabel:
    def __init__(self):
        self.text = ""
    def config(self, text):
        self.text = text
    def grid(self):
        pass

# Fixture for temporary directory to hold Excel files.
@pytest.fixture
def temp_excel_dir():
    dirpath = tempfile.mkdtemp()
    yield dirpath
    shutil.rmtree(dirpath)

# Helper to create a dummy Excel file with openpyxl.
def create_dummy_excel(file_path, sheet_name="Sheet1", cell_data=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if cell_data:
        for cell, value in cell_data.items():
            ws[cell] = value
    wb.save(file_path)

def test_fetch_latest_version(monkeypatch):
    # Simulate a successful GitHub response.
    class DummyResponse:
        def __init__(self):
            self._json = {"tag_name": "v1.2.3"}
        def raise_for_status(self):
            pass
        def json(self):
            return self._json
    monkeypatch.setattr(requests, "get", lambda url, timeout: DummyResponse())
    version = fetch_latest_version()
    assert version == "v1.2.3"

def test_clean_pipes():
    assert clean_pipes("|   |Hello|  |World|") == "Hello|World|"
    assert clean_pipes("|a=1||b=2|") == "a=1|b=2|"
    assert clean_pipes("|   |") == ""

def test_remove_key_value_pair_from_cell():
    cell_value = "a=1|b=2|c=3|"
    assert remove_key_value_pair_from_cell(cell_value, "b") == "a=1|c=3"
    # If key not found, returns original
    assert remove_key_value_pair_from_cell(cell_value, "d") == "a=1|b=2|c=3"


def test_search_replace_or_remove_key(temp_excel_dir):
    # Create a dummy Excel file with known content.
    file_path = os.path.join(temp_excel_dir, "test.xlsx")
    create_dummy_excel(file_path, cell_data={"A1": "a=1|b=2|c=3|"})
    log = DummyLog()
    key_found = [False]
    search_replace_or_remove_key(file_path, "b", "42", False, log, key_found)
    wb = openpyxl.load_workbook(file_path)
    val = wb.active["A1"].value
    # clean_pipes removes trailing pipes: expected "a=1|b=42|c=3"
    assert val == "a=1|b=42|c=3|"

def test_backup_excel_files(temp_excel_dir):
    # Create a dummy Excel file in source.
    source_file = os.path.join(temp_excel_dir, "dummy.xlsx")
    create_dummy_excel(source_file, cell_data={"A1": "dummy content"})
    backup_dir = os.path.join(temp_excel_dir, "Backup")
    log = DummyLog()
    backup_excel_files(temp_excel_dir, backup_dir, log)
    backup_file = os.path.join(backup_dir, "dummy.xlsx")
    assert os.path.exists(backup_file)
    combined_log = " ".join(log.messages).lower()
    assert "backup directory" in combined_log


def test_process_excel_files(temp_excel_dir):
    # Create a dummy Excel file with content that will be processed.
    file_path = os.path.join(temp_excel_dir, "test.xlsx")
    create_dummy_excel(file_path, cell_data={"A1": "a=1|b=2|c=3|"})

    # Define backup directory inside temp_excel_dir.
    backup_dir = os.path.join(temp_excel_dir, "Backup")
    dummy_log = DummyLog()

    # First, create the backup.
    backup_excel_files(temp_excel_dir, backup_dir, dummy_log)

    # Now, perform processing on the originals.
    prog = DummyProgressBar()
    percent = DummyLabel()
    process_excel_files(temp_excel_dir, backup_dir, "b", "100", False, dummy_log, prog, percent)

    # Check that the original file was modified.
    wb = openpyxl.load_workbook(file_path)
    val = wb.active["A1"].value
    assert val == "a=1|b=100|c=3|"

    # Check that the backup file remains unchanged.
    backup_file = os.path.join(backup_dir, "test.xlsx")
    wb_backup = openpyxl.load_workbook(backup_file)
    val_backup = wb_backup.active["A1"].value
    assert val_backup == "a=1|b=2|c=3|"


def test_process_value_cells(temp_excel_dir):
    file_path = os.path.join(temp_excel_dir, "test.xlsx")
    create_dummy_excel(file_path, cell_data={"A1": "Hello World", "B1": "Foo Bar"})
    log = DummyLog()
    process_value_cells(file_path, "World", "Universe", log)
    wb = openpyxl.load_workbook(file_path)
    val = wb.active["A1"].value
    assert val == "Hello Universe"
    # Test that non-existent string leaves file unchanged.
    process_value_cells(file_path, "Nonexistent", "Test", log)
    wb = openpyxl.load_workbook(file_path)
    assert wb.active["A1"].value == "Hello Universe"

def test_process_value_in_directory(temp_excel_dir):
    file1 = os.path.join(temp_excel_dir, "file1.xlsx")
    file2 = os.path.join(temp_excel_dir, "file2.xlsx")
    create_dummy_excel(file1, cell_data={"A1": "Alpha Beta", "B1": "Gamma"})
    create_dummy_excel(file2, cell_data={"A1": "Beta Gamma", "B1": "Delta"})
    log = DummyLog()
    prog = DummyProgressBar()
    percent = DummyLabel()
    process_value_in_directory(temp_excel_dir, "Beta", "Replaced", log, prog, percent, backup_dir=None)
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)
    assert "Replaced" in wb1.active["A1"].value
    assert "Replaced" in wb2.active["A1"].value

def test_start_value_replacement(temp_excel_dir):
    file_path = os.path.join(temp_excel_dir, "test.xlsx")
    create_dummy_excel(file_path, cell_data={"A1": "Hello World", "B1": "Test"})
    log = DummyLog()
    prog = DummyProgressBar()
    percent = DummyLabel()
    # Call start_value_replacement with an empty backup_dir_value to trigger default behavior.
    start_value_replacement(temp_excel_dir, "World", "Universe", log, prog, percent, backup_dir_value="")
    backup_dir = os.path.join(temp_excel_dir, "Backup")
    assert os.path.exists(backup_dir)
    wb = openpyxl.load_workbook(file_path)
    assert wb.active["A1"].value == "Hello Universe"
    # Check backup file remains with original content.
    backup_file = os.path.join(backup_dir, "test.xlsx")
    wb_backup = openpyxl.load_workbook(backup_file)
    assert wb_backup.active["A1"].value == "Hello World"

def test_open_github_link(monkeypatch):
    called = False
    def dummy_open(url):
        nonlocal called
        called = True
        assert "github.com" in url
    monkeypatch.setattr("functions.webbrowser.open", dummy_open)
    from functions import open_github_link
    open_github_link()
    assert called


def test_check_for_updates(monkeypatch):
    from tkinter import messagebox
    # Override showinfo so that it doesn't try to create a GUI.
    monkeypatch.setattr(messagebox, "showinfo", lambda title, message, **kwargs: None)

    def dummy_fetch():
        return "v999.0"

    monkeypatch.setattr("functions.fetch_latest_version", dummy_fetch)
    from functions import check_for_updates
    # This call should now not raise a TclError.
    check_for_updates()

if __name__ == "__main__":
    pytest.main()

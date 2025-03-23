import openpyxl
import os
import re
import requests
import webbrowser
import threading
import shutil
from tkinter import Tk, Label, Entry, Button, filedialog, Text, Scrollbar, VERTICAL, END, IntVar, Checkbutton, \
    messagebox, Toplevel
from tkinter.ttk import Progressbar

# GitHub Repository Info
GITHUB_REPO = "Maksymilianx/Excel_word_changer"
FALLBACK_VERSION = "1.0.0"


def fetch_latest_version():
    """Fetch the latest release tag from GitHub."""
    url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
    try:
        response = requests.get(url, timeout=5)
        response.raise_for_status()
        latest_version = response.json().get("tag_name", FALLBACK_VERSION)
        return latest_version
    except requests.RequestException:
        return FALLBACK_VERSION


VERSION = fetch_latest_version()


def backup_excel_files(source_dir, backup_dir, log_widget):
    """
    Copy all .xlsx files from source_dir (including subfolders) to backup_dir,
    preserving the folder structure.
    """
    for root, _, files in os.walk(source_dir):
        # Skip the backup folder if it's inside the processing directory
        if os.path.abspath(root) == os.path.abspath(backup_dir):
            continue
        for file in files:
            if file.endswith('.xlsx'):
                source_file = os.path.join(root, file)
                rel_path = os.path.relpath(root, source_dir)
                target_folder = os.path.join(backup_dir, rel_path)
                os.makedirs(target_folder, exist_ok=True)
                target_file = os.path.join(target_folder, file)
                shutil.copy2(source_file, target_file)
                log_widget.insert(END, f"Backup: {source_file} -> {target_file}\n", "info")


def remove_key_value_pair_from_cell(cell_value, key):
    """
    Remove the key-value pair and the surrounding pipes from the cell value.
    Returns the updated cell value if a change is made, else None.
    """
    pattern = r'\|?' + re.escape(key) + r'=[^|]*\|?'
    new_value = re.sub(pattern, '|', cell_value)
    # Clean up duplicate pipes and trim leading/trailing pipes
    new_value = re.sub(r'\|{2,}', '|', new_value).strip('|')
    return new_value if new_value != cell_value else None


def search_replace_or_remove_key(file_path, key, new_value, remove_key, log_widget, key_found):
    """Search for a key in an Excel file, replace its value, or remove it entirely."""
    try:
        workbook = openpyxl.load_workbook(file_path)
        found = False
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if remove_key:
                            updated_cell = remove_key_value_pair_from_cell(cell.value, key)
                        else:
                            pattern = rf"\|?{key}=[^|]*\|?"
                            updated_cell = re.sub(pattern, f"|{key}={new_value}|", cell.value)
                        if updated_cell and updated_cell != cell.value:
                            found = True
                            key_found[0] = True
                            cell.value = updated_cell
        if found:
            workbook.save(file_path)
            log_widget.insert(END, f"✅ Updated: {file_path}\n", "success")
    except Exception as e:
        log_widget.insert(END, f"❌ Error processing {file_path}: {e}\n", "error")


def process_excel_files(directory_path, backup_dir, key, new_value, remove_key, log_widget, progress_bar):
    """Process all Excel files in the directory (excluding the backup folder), updating the progress bar."""
    log_widget.delete(1.0, END)
    log_widget.insert(END, f"🔄 Processing files in {directory_path}...\n", "info")

    # Count total Excel files (excluding those in the backup folder)
    total_files = 0
    for root, dirs, files in os.walk(directory_path):
        if backup_dir:
            backup_basename = os.path.basename(backup_dir)
            if backup_basename in dirs:
                dirs.remove(backup_basename)
        for file in files:
            if file.endswith('.xlsx'):
                total_files += 1

    if total_files == 0:
        log_widget.insert(END, "❌ No Excel files found.\n", "error")
        progress_bar["value"] = 0
        return

    progress_bar["maximum"] = total_files
    progress_bar["value"] = 0
    key_found = [False]
    processed = 0
    try:
        for root, dirs, files in os.walk(directory_path):
            if backup_dir:
                backup_basename = os.path.basename(backup_dir)
                if backup_basename in dirs:
                    dirs.remove(backup_basename)
            for file in files:
                if file.endswith('.xlsx'):
                    file_path = os.path.join(root, file)
                    search_replace_or_remove_key(file_path, key, new_value, remove_key, log_widget, key_found)
                    processed += 1
                    progress_bar["value"] = processed
                    progress_bar.update_idletasks()
        if not key_found[0]:
            log_widget.insert(END, f"⚠ Warning: The key '{key}' was not found in any file.\n", "warning")
            show_custom_warning_popup(f"The key '{key}' was not found in any file.")
        log_widget.insert(END, "✅ Process completed!\n", "success")
    except Exception as e:
        log_widget.insert(END, f"❌ An error occurred: {e}\n", "error")


def show_custom_warning_popup(message):
    """Custom warning popup, centered on the screen."""
    popup = Toplevel()
    popup.title("Warning")
    popup.geometry("300x100")
    popup.update_idletasks()
    x_position = (popup.winfo_screenwidth() // 2) - (300 // 2)
    y_position = (popup.winfo_screenheight() // 2) - (100 // 2)
    popup.geometry(f"300x100+{x_position}+{y_position}")
    label = Label(popup, text=message, fg="red", font=("Arial", 12, "bold"))
    label.pack(pady=10)
    close_button = Button(popup, text="OK", command=popup.destroy)
    close_button.pack(pady=5)
    popup.grab_set()


def check_for_updates():
    """Check if a newer version is available and show a popup."""
    latest_version = fetch_latest_version()
    if latest_version > VERSION:
        messagebox.showinfo("Update Available",
                            f"A new version ({latest_version}) is available!\nVisit GitHub to download.")
    else:
        messagebox.showinfo("Up to Date", "You are using the latest version.")


def browse_directory(entry_widget):
    """Open a directory selection dialog."""
    directory = filedialog.askdirectory()
    if directory:
        entry_widget.delete(0, END)
        entry_widget.insert(0, directory)


def open_github_link():
    """Open GitHub repository in a browser."""
    webbrowser.open("https://github.com/Maksymilianx/Excel_word_changer")


def start_processing(directory_entry, backup_entry, key_entry, value_entry, remove_key_var, log_widget, progress_bar):
    """
    Validate input, perform backup, and start processing files in a separate thread.
    If the backup directory field is empty, a folder named "Backup" is created inside the processing directory.
    If the user selects the processing directory as the backup directory, a subfolder "Backup" inside the processing
    directory is used instead.
    """
    directory = directory_entry.get()
    backup_dir = backup_entry.get()
    key = key_entry.get()
    new_value = value_entry.get()
    remove_key = remove_key_var.get()

    if not directory or not os.path.exists(directory):
        log_widget.insert(END, "❌ Please select a valid processing directory.\n", "error")
        return

    # If backup directory is empty, default to a folder named "Backup" inside the processing directory.
    if not backup_dir:
        backup_dir = os.path.join(directory, "Backup")
        log_widget.insert(END,
                          "ℹ No backup directory specified. Using default 'Backup' folder inside the processing directory.\n",
                          "info")
    # If the backup directory is the same as the processing directory, automatically use a subfolder named "Backup".
    elif os.path.abspath(backup_dir) == os.path.abspath(directory):
        log_widget.insert(END,
                          "⚠ Backup directory is the same as processing directory. Using a subfolder 'Backup' instead.\n",
                          "warning")
        backup_dir = os.path.join(directory, "Backup")

    if not key:
        log_widget.insert(END, "❌ Please enter a key to search for.\n", "error")
        return
    if not remove_key and not new_value:
        log_widget.insert(END, "❌ Please enter a new value or check the 'Remove Key' option.\n", "error")
        return
    if remove_key:
        confirm = messagebox.askyesno("Confirm Deletion", f"Are you sure you want to remove '{key}' and its value?")
        if not confirm:
            log_widget.insert(END, "⚠ Deletion canceled by user.\n", "warning")
            return

    # Backup the original Excel files before processing
    log_widget.insert(END, "🔄 Backing up files...\n", "info")
    backup_excel_files(directory, backup_dir, log_widget)
    log_widget.insert(END, "✅ Backup completed!\n", "success")

    # Run the processing in a separate thread to keep the UI responsive.
    threading.Thread(target=process_excel_files,
                     args=(directory, backup_dir, key, new_value, remove_key, log_widget, progress_bar)).start()


# ---------------------- GUI Setup ---------------------- #

# Create main window and hide it initially for splash screen
root = Tk()
root.withdraw()  # Hide the main window while splash is active
root.title(f"xlxs fixer - v{VERSION}")

# Create a splash screen
splash = Toplevel()
splash.overrideredirect(True)  # Remove window decorations
splash_label = Label(splash, text="Loading xlxs fixer...", font=("Helvetica", 18), bg="white", fg="black")
splash_label.pack(expand=True, fill="both")
splash.update_idletasks()
splash_width = 300
splash_height = 200
x_position = (splash.winfo_screenwidth() - splash_width) // 2
y_position = (splash.winfo_screenheight() - splash_height) // 2
splash.geometry(f"{splash_width}x{splash_height}+{x_position}+{y_position}")


# After 3 seconds, destroy splash and show main window
def close_splash():
    splash.destroy()
    root.deiconify()  # Show the main window


splash.after(3000, close_splash)

# Backup Directory selection (Row 0)
Label(root, text="Select Backup Directory:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
backup_entry = Entry(root, width=50)
backup_entry.grid(row=0, column=1, padx=10, pady=5)
backup_browse_button = Button(root, text="Browse", command=lambda: browse_directory(backup_entry))
backup_browse_button.grid(row=0, column=2, padx=10, pady=5)

# Processing Directory selection (Row 1)
Label(root, text="Select Processing Directory:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
directory_entry = Entry(root, width=50)
directory_entry.grid(row=1, column=1, padx=10, pady=5)
directory_browse_button = Button(root, text="Browse", command=lambda: browse_directory(directory_entry))
directory_browse_button.grid(row=1, column=2, padx=10, pady=5)

# Key input (Row 2)
Label(root, text="Key to Search:").grid(row=2, column=0, padx=10, pady=5, sticky="w")
key_entry = Entry(root, width=50)
key_entry.grid(row=2, column=1, padx=10, pady=5)

# New value input (Row 3)
Label(root, text="New Value:").grid(row=3, column=0, padx=10, pady=5, sticky="w")
value_entry = Entry(root, width=50)
value_entry.grid(row=3, column=1, padx=10, pady=5)

# Remove Key Checkbox (Row 4)
remove_key_var = IntVar()
remove_key_checkbox = Checkbutton(root, text="Remove Key", variable=remove_key_var)
remove_key_checkbox.grid(row=4, column=1, pady=5)

# Log display (Row 5)
log_widget = Text(root, height=15, width=70)
log_widget.grid(row=5, column=0, columnspan=3, padx=10, pady=10)
scrollbar = Scrollbar(root, orient=VERTICAL, command=log_widget.yview)
scrollbar.grid(row=5, column=3, sticky="ns")
log_widget.config(yscrollcommand=scrollbar.set)
log_widget.tag_config("error", foreground="red")
log_widget.tag_config("warning", foreground="orange")
log_widget.tag_config("success", foreground="green")
log_widget.tag_config("info", foreground="blue")
log_widget.insert(END, f"🛠 Version: {VERSION}\n", "info")

# Progress Bar (Row 6)
progress_bar = Progressbar(root, orient="horizontal", mode="determinate", length=400)
progress_bar.grid(row=6, column=0, columnspan=3, padx=10, pady=10)

# Start Processing Button (Row 7)
Button(root, text="Start Processing",
       command=lambda: start_processing(directory_entry, backup_entry, key_entry, value_entry, remove_key_var,
                                        log_widget, progress_bar)).grid(row=7, column=1, pady=10)

# Check for Updates button (Row 8)
Button(root, text="Check for Updates", command=check_for_updates).grid(row=8, column=1, pady=5)

# GitHub link (Row 9)
github_label = Label(root, text="View on GitHub", fg="blue", cursor="hand2")
github_label.grid(row=9, column=1, pady=10)
github_label.bind("<Button-1>", lambda e: open_github_link())

root.mainloop()
